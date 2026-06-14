from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
import io


# Color constants — reference-doc style (yellow/beige palette)
HEADER_YELLOW = "FFFFCC"   # light yellow for label/header cells
SECTION_YELLOW = "FFFACD"  # slightly warmer yellow for section rows
MARK_YELLOW = "FFFF00"     # highlight for △ rows
RED = "FF0000"
WHITE = "FFFFFF"
GRAY = "D9D9D9"
BLACK = "000000"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color=BLACK, size=11) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_cell(ws, row, col, value, fill=None, font=None, alignment=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = _border()
    return cell


def generate_report_excel(report_data: dict, photos: list, output_path: Path):
    wb = Workbook()

    _build_report_sheet(wb, report_data)
    _build_photo_sheet(wb, photos, report_data.get("photo_descriptions", []))

    wb.save(str(output_path))


def _build_report_sheet(wb: Workbook, data: dict):
    ws = wb.active
    ws.title = "点検報告書"

    # Column widths: A=label, B-D=value, E=code
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14

    # Dropdown validators (created once, applied per cell)
    dv_method = DataValidation(
        type="list",
        formula1='"A,B,C,AC,AB,BC,ABC"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_judgment = DataValidation(
        type="list",
        formula1='"○,△,ー"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_method)
    ws.add_data_validation(dv_judgment)

    row = 1

    # JIO label (small text above title, matching reference)
    ws.merge_cells(f"A{row}:E{row}")
    jio_cell = ws.cell(row=row, column=1, value="JIO（日本住宅保証検査機構）")
    jio_cell.font = _font(size=9)
    jio_cell.alignment = _left()
    ws.row_dimensions[row].height = 14
    row += 1

    # Title row — spaced characters, light yellow background
    ws.merge_cells(f"A{row}:E{row}")
    title_cell = ws.cell(row=row, column=1,
                         value="ア フ タ ー ハ ウ ス 点 検 報 告 書")
    title_cell.fill = _fill(HEADER_YELLOW)
    title_cell.font = _font(bold=True, size=16)
    title_cell.alignment = _center()
    title_cell.border = _border()
    ws.row_dimensions[row].height = 36
    row += 1

    row += 1  # blank row

    # Basic info block — label | value (B-D merged) | code (E)
    # 登録物件名 and 事業者名 have codes; others span B-E
    info_with_code = [
        ("登録物件名", data.get("property_name", ""), data.get("property_no", "")),
        ("事業者名",   data.get("company_name", ""),  data.get("company_no", "")),
    ]
    info_no_code = [
        ("引渡し年月", data.get("delivery_date", "")),
        ("実施日",     data.get("inspection_date", "")),
        ("報告者",     data.get("inspector", "")),
    ]

    for label, value, code in info_with_code:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = _font(bold=True)
        label_cell.alignment = _left()
        label_cell.border = _border()

        ws.merge_cells(f"B{row}:D{row}")
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = _font()
        val_cell.alignment = _left()
        val_cell.border = _border()

        code_cell = ws.cell(row=row, column=5, value=code)
        code_cell.font = _font()
        code_cell.alignment = _center()
        code_cell.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

    for label, value in info_no_code:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = _font(bold=True)
        label_cell.alignment = _left()
        label_cell.border = _border()

        ws.merge_cells(f"B{row}:E{row}")
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = _font()
        val_cell.alignment = _left()
        val_cell.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # blank row

    # Legend
    ws.merge_cells(f"A{row}:E{row}")
    legend_cell = ws.cell(
        row=row,
        column=1,
        value='【凡例】「方法」A：目視確認 B：動作確認 C：聴取確認　「判定」○：事象なし △：事象あり ー：該当なし',
    )
    legend_cell.font = _font(size=10)
    legend_cell.alignment = _left()
    ws.row_dimensions[row].height = 14
    row += 1

    row += 1  # blank row

    # Sections
    for section in data.get("sections", []):
        sec_no = section.get("section_no", "")
        sec_name = section.get("section_name", "")

        # Section header — bold black text, light yellow background
        ws.merge_cells(f"A{row}:E{row}")
        sec_cell = ws.cell(row=row, column=1, value=f"{sec_no}. {sec_name}")
        sec_cell.fill = _fill(SECTION_YELLOW)
        sec_cell.font = _font(bold=True)
        sec_cell.alignment = _left()
        sec_cell.border = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        # Column header
        headers = ["No.", "項　目", "方 法", "判 定", "備考"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = _fill(HEADER_YELLOW)
            cell.font = _font(bold=True)
            cell.alignment = _center()
            cell.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        # Items
        for item in section.get("items", []):
            judgment = item.get("judgment", "○")
            is_bad = judgment in ("△", "×")

            row_fill = _fill(MARK_YELLOW) if is_bad else None

            # No.
            no_cell = ws.cell(row=row, column=1, value=item.get("no", ""))
            if row_fill:
                no_cell.fill = row_fill
            no_cell.font = _font(bold=False)
            no_cell.alignment = _center()
            no_cell.border = _border()

            # Description
            desc_cell = ws.cell(row=row, column=2, value=item.get("description", ""))
            if row_fill:
                desc_cell.fill = row_fill
            desc_cell.font = _font()
            desc_cell.alignment = _left()
            desc_cell.border = _border()

            # Method — dropdown
            method_cell = ws.cell(row=row, column=3, value=item.get("method", "AC"))
            if row_fill:
                method_cell.fill = row_fill
            method_cell.font = _font()
            method_cell.alignment = _center()
            method_cell.border = _border()
            dv_method.add(method_cell)

            # Judgment — dropdown
            judg_cell = ws.cell(row=row, column=4, value=judgment)
            if row_fill:
                judg_cell.fill = row_fill
            judg_cell.font = _font(bold=is_bad, color=RED if is_bad else BLACK)
            judg_cell.alignment = _center()
            judg_cell.border = _border()
            dv_judgment.add(judg_cell)

            # Notes
            notes_cell = ws.cell(row=row, column=5, value=item.get("notes", ""))
            if row_fill:
                notes_cell.fill = row_fill
            notes_cell.font = _font()
            notes_cell.alignment = _left()
            notes_cell.border = _border()

            ws.row_dimensions[row].height = 18
            row += 1

        # 備考 row per section (matching reference doc style)
        ws.merge_cells(f"A{row}:E{row}")
        biko_cell = ws.cell(row=row, column=1, value="【備考】")
        biko_cell.fill = _fill(HEADER_YELLOW)
        biko_cell.font = _font(bold=True)
        biko_cell.alignment = _left()
        biko_cell.border = _border()
        ws.row_dimensions[row].height = 24
        row += 1

        row += 1  # spacing between sections

    # Special notes
    special_notes = data.get("special_notes", [])
    if special_notes:
        ws.merge_cells(f"A{row}:E{row}")
        note_header = ws.cell(row=row, column=1, value="【特記事項】")
        note_header.fill = _fill(HEADER_YELLOW)
        note_header.font = _font(bold=True)
        note_header.alignment = _left()
        note_header.border = _border()
        row += 1

        for note in special_notes:
            ws.merge_cells(f"A{row}:E{row}")
            note_cell = ws.cell(row=row, column=1, value=f"・{note}")
            note_cell.font = _font()
            note_cell.alignment = _left()
            note_cell.border = _border()
            ws.row_dimensions[row].height = 20
            row += 1


def _build_photo_sheet(wb: Workbook, photos: list, photo_descriptions: list):
    ws = wb.create_sheet(title="現場写真")

    # Col A: photo, Col B: No./filename, Col C: AI description, Col D: remarks
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 35

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="現場写真一覧")
    title_cell.fill = _fill(HEADER_YELLOW)
    title_cell.font = _font(bold=True, size=13)
    title_cell.alignment = _center()
    title_cell.border = _border()
    ws.row_dimensions[1].height = 28

    # Column headers
    col_headers = ["写真", "No. / ファイル名", "AI解析内容", "備考"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = _fill(HEADER_YELLOW)
        cell.font = _font(bold=True)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[2].height = 18

    # Build filename → description map
    desc_map = {pd.get("filename", ""): pd.get("description", "") for pd in photo_descriptions}

    row = 3
    for i, photo_path in enumerate(photos):
        # Try to embed image
        try:
            from PIL import Image as PILImage

            with PILImage.open(str(photo_path)) as pil_img:
                pil_img.thumbnail((280, 200), PILImage.LANCZOS)
                img_io = io.BytesIO()
                fmt = pil_img.format or "JPEG"
                if fmt not in ("JPEG", "PNG", "GIF", "BMP"):
                    fmt = "PNG"
                pil_img.save(img_io, format=fmt)
                img_io.seek(0)

            xl_img = XLImage(img_io)
            xl_img.width = 280
            xl_img.height = 200

            ws.add_image(xl_img, f"A{row}")
            ws.row_dimensions[row].height = 155
        except Exception:
            ws.cell(row=row, column=1, value=f"[画像: {photo_path.name}]").border = _border()
            ws.row_dimensions[row].height = 155

        # No. / filename
        no_cell = ws.cell(row=row, column=2, value=f"No.{i + 1}\n{photo_path.name}")
        no_cell.alignment = _left()
        no_cell.border = _border()
        no_cell.font = _font(size=10)

        # AI description
        filename = photo_path.name
        description = desc_map.get(filename, "")
        desc_cell = ws.cell(row=row, column=3, value=description)
        desc_cell.alignment = _left()
        desc_cell.border = _border()
        desc_cell.font = _font(size=10)

        # Remarks (blank for manual input)
        remarks_cell = ws.cell(row=row, column=4, value="")
        remarks_cell.alignment = _left()
        remarks_cell.border = _border()

        row += 1

    # If no photos
    if not photos:
        ws.merge_cells("A3:D3")
        ws.cell(row=3, column=1, value="（写真なし）").alignment = _center()
